"""
utils/excel_processor.py
=========================
Builds the professional Excel database from scraped records.
3 sheets: Full Database, Dashboard, Summons Ready
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Palette ───────────────────────────────────────────────────────────────────
NAVY   = "0A1628"
GOLD   = "C9A84C"
LIGHT  = "F8FAFC"
WHITE  = "FFFFFF"
RED    = "DC2626"
ORANGE = "EA580C"
AMBER  = "D97706"
GREEN  = "059669"
SLATE  = "64748B"

URGENCY_COLORS = {
    "CRITICAL": RED,
    "HIGH":     ORANGE,
    "MEDIUM":   AMBER,
    "LOW":      GREEN,
}

_thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr_cell(cell, text, fill_color=NAVY):
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _body_cell(cell, value, row_idx, bold=False, color=None):
    cell.value = str(value) if value else ""
    cell.font = Font(name="Calibri", bold=bold, size=9,
                     color=color or ("1E293B" if not bold else NAVY))
    cell.fill = PatternFill("solid", fgColor=LIGHT if row_idx % 2 == 0 else WHITE)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER


def build_excel(records: list) -> bytes:
    wb = Workbook()

    # ══════════════════════════════════════════════════════
    # SHEET 1 — Full Case Database
    # ══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📋 Case Database"

    headers = [
        "Date Extracted", "Document Type", "Urgency",
        "Tenant Name", "Unit", "Full Address", "Email", "Phone",
        "Landlord / LLC", "Case Number", "Agency",
        "Violation Type", "Violation Class",
        "Amount Owed ($)", "Monthly Rent ($)", "Months Overdue", "Total Overdue ($)",
        "Issue Date", "Deadline Date", "Hearing Date",
        "Lease Start", "Lease End",
        "Summary", "Recommended Action", "Summons Type", "Status"
    ]
    col_widths = [
        18, 20, 11,
        24, 7, 36, 26, 15,
        26, 18, 9,
        26, 14,
        14, 14, 14, 14,
        14, 14, 14,
        12, 12,
        50, 38, 24, 12
    ]

    ws1.row_dimensions[1].height = 38
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr_cell(ws1.cell(1, ci), h)
        ws1.column_dimensions[ws1.cell(1, ci).column_letter].width = w

    for ri, rec in enumerate(records, 2):
        urg = rec.get("urgency", "MEDIUM")
        addr = " ".join(filter(None, [
            rec.get("tenant_address", ""),
            rec.get("tenant_city", ""),
            rec.get("tenant_state", ""),
            rec.get("tenant_zip", ""),
        ]))
        row_data = [
            rec.get("extracted_date", ""),
            rec.get("document_type", ""),
            urg,
            rec.get("tenant_name", ""),
            rec.get("tenant_unit", ""),
            addr,
            rec.get("tenant_email", ""),
            rec.get("tenant_phone", ""),
            rec.get("landlord_name", ""),
            rec.get("case_number", ""),
            rec.get("agency", ""),
            rec.get("violation_type", ""),
            rec.get("violation_class", ""),
            rec.get("amount_owed", ""),
            rec.get("monthly_rent", ""),
            rec.get("months_overdue", ""),
            rec.get("total_overdue", ""),
            rec.get("issue_date", ""),
            rec.get("deadline_date", ""),
            rec.get("hearing_date", ""),
            rec.get("lease_start", ""),
            rec.get("lease_end", ""),
            rec.get("issue_summary", ""),
            rec.get("recommended_action", ""),
            rec.get("summons_type", ""),
            "Pending",
        ]
        ws1.row_dimensions[ri].height = 26
        for ci, val in enumerate(row_data, 1):
            cell = ws1.cell(ri, ci)
            if ci == 3:  # Urgency column — colored
                cell.value = str(val)
                cell.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
                cell.fill = PatternFill("solid", fgColor=URGENCY_COLORS.get(urg, AMBER))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER
            else:
                _body_cell(cell, val, ri)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ══════════════════════════════════════════════════════
    # SHEET 2 — Dashboard
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("📊 Dashboard")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 22

    def dash_section(row, text):
        c = ws2.cell(row, 1, text)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[row].height = 30
        ws2.merge_cells(f"A{row}:B{row}")

    def dash_kv(row, label, value, bold_val=False, val_color=None):
        c1 = ws2.cell(row, 1, label)
        c1.font = Font(name="Calibri", size=10)
        c1.alignment = Alignment(indent=3, vertical="center")
        c1.border = BORDER
        c1.fill = PatternFill("solid", fgColor=LIGHT if row % 2 == 0 else WHITE)
        c2 = ws2.cell(row, 2, value)
        c2.font = Font(name="Calibri", bold=bold_val, size=10, color=val_color or (NAVY if bold_val else SLATE))
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = BORDER
        c2.fill = PatternFill("solid", fgColor=LIGHT if row % 2 == 0 else WHITE)
        ws2.row_dimensions[row].height = 24

    total = len(records)
    critical = sum(1 for r in records if r.get("urgency") == "CRITICAL")
    high = sum(1 for r in records if r.get("urgency") == "HIGH")
    summons_n = sum(1 for r in records if r.get("summons_applicable") in [True, "true", "True"])
    total_owed = 0
    for r in records:
        try:
            total_owed += float(str(r.get("total_overdue") or r.get("amount_owed") or "0")
                                .replace(",", "").replace("$", "") or 0)
        except Exception:
            pass

    dash_section(1, f"🏛️  LANDLORDAI — CASE SUMMARY  |  Generated {datetime.now().strftime('%b %d, %Y %H:%M')}")
    ws2.row_dimensions[2].height = 6
    dash_section(3, "📊  OVERVIEW")
    dash_kv(4,  "Total Cases in Database",    total,                    True)
    dash_kv(5,  "🔴  Critical Urgency Cases", critical,                 True, RED)
    dash_kv(6,  "🟠  High Urgency Cases",     high,                     False, ORANGE)
    dash_kv(7,  "⚖️   Summons Applicable",    summons_n,                True)
    dash_kv(8,  "💰  Total Overdue Rent",      f"${total_owed:,.2f}",   True, RED)
    ws2.row_dimensions[9].height = 6
    dash_section(10, "⏰  UPCOMING DEADLINES")
    r_n = 11
    for rec in records:
        dd = rec.get("deadline_date", "")
        if dd and dd not in ("null", "N/A", "None", ""):
            dash_kv(r_n, rec.get("tenant_name", "Unknown"), f"Deadline: {dd}")
            r_n += 1
    if r_n == 11:
        dash_kv(11, "No upcoming deadlines found", "—")

    # ══════════════════════════════════════════════════════
    # SHEET 3 — Summons Ready
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("⚖️ Summons Ready")
    s_headers = ["Tenant Name", "Unit", "Address", "Summons Type", "Amount Due ($)", "Deadline", "Status"]
    s_widths   = [26, 8, 38, 28, 16, 16, 14]
    ws3.row_dimensions[1].height = 32
    for ci, (h, w) in enumerate(zip(s_headers, s_widths), 1):
        _hdr_cell(ws3.cell(1, ci), h)
        ws3.column_dimensions[ws3.cell(1, ci).column_letter].width = w

    r3 = 2
    for rec in records:
        if rec.get("summons_applicable") in [True, "true", "True"]:
            addr = f"{rec.get('tenant_address','')} {rec.get('tenant_city','')}".strip()
            amount = rec.get("total_overdue") or rec.get("amount_owed") or ""
            row3 = [
                rec.get("tenant_name", ""),
                rec.get("tenant_unit", ""),
                addr,
                rec.get("summons_type", ""),
                amount,
                rec.get("deadline_date", ""),
                "⚡ Ready",
            ]
            ws3.row_dimensions[r3].height = 24
            for ci, val in enumerate(row3, 1):
                _body_cell(ws3.cell(r3, ci), val, r3)
            r3 += 1

    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
